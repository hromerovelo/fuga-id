"""
BSD 2-Clause License

Copyright (c) 2024, Hilda Romero-Velo
All rights reserved.

Redistribution and use in source and binary forms, with or without
modification, are permitted provided that the following conditions are met:

* Redistributions of source code must retain the above copyright notice, this
  list of conditions and the following disclaimer.

* Redistributions in binary form must reproduce the above copyright notice, this
  list of conditions and the following disclaimer in the documentation and/or
  other materials provided with the distribution.

THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE
FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR
SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
"""

"""
  Created by Hilda Romero-Velo on December 2024.
"""

""" 
  This script compiles statistics retrieved from que folkoteca.db and returns
  an .xlsx file. 
"""

import os
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font
import shutil
import argparse


def execute_query(conn, query):
    """Execute a SQL query and return the result as a DataFrame."""
    return pd.read_sql_query(query, conn)


def copy_tables(src_conn, dest_conn, tables):
    """Copy specified tables from source database to destination database."""
    for table in tables:
        dest_conn.execute(f"DROP TABLE IF EXISTS {table}")
        src_df = pd.read_sql_query(f"SELECT * FROM {table}", src_conn)
        src_df.to_sql(table, dest_conn, if_exists="append", index=False)


def prepare_database(db_path, global_db_path, use_global_alignment):
    """Prepare the database and return the connection and queries module."""
    if use_global_alignment:
        # Copy global_folkoteca.db to /analysis_results/
        script_dir = os.path.dirname(os.path.abspath(__file__))
        analysis_dir = os.path.join(script_dir, "analysis_results")
        os.makedirs(analysis_dir, exist_ok=True)
        target_db_path = os.path.join(analysis_dir, "analysis_folkoteca.db")
        shutil.copyfile(global_db_path, target_db_path)

        conn = sqlite3.connect(target_db_path)
        # Copy tables to analysis_folkoteca.db
        db_conn = sqlite3.connect(db_path)
        tables_to_copy = ["recording", "query", "search", "search_results"]
        copy_tables(db_conn, conn, tables_to_copy)
        db_conn.close()

        # Use global alignment queries
        import report_queries_global_hits as queries_module
    else:
        conn = sqlite3.connect(db_path)
        import report_queries as queries_module

    return conn, queries_module


def calculate_metrics(db_path, global_db_path, output_file, use_global_alignment):
    conn, queries_module = prepare_database(
        db_path, global_db_path, use_global_alignment
    )

    queries = {
        "general_metrics": queries_module.general_metrics,
        "metrics_by_instrument": queries_module.metrics_by_instrument,
        "metrics_by_musical_form": queries_module.metrics_by_musical_form,
        "metrics_by_query_sequence": queries_module.metrics_by_query_sequence,
        "metrics_by_score": queries_module.metrics_by_score,
    }
    wb = Workbook()

    for sheet_name, query_set in queries.items():
        # Execute all queries in the queries module
        num_queries_df = execute_query(conn, query_set["num_queries"])
        if sheet_name == "general_metrics":
            avg_times_df = execute_query(conn, query_set["avg_times"])
        count_top_1_df = execute_query(conn, query_set["count_top_1"])
        count_top_3_df = execute_query(conn, query_set["count_top_3"])
        count_top_5_df = execute_query(conn, query_set["count_top_5"])
        mrr_sum_df = execute_query(conn, query_set["mrr_sum"])

        # Merge DataFrames on relevant columns
        merge_columns = ["algorithm", "search_type"]
        if sheet_name != "general_metrics":
            split_name = sheet_name.split("_by_")[1]
            if split_name == "query_sequence":
                merge_columns.insert(0, "sequence_length")
            else:
                merge_columns.insert(0, split_name)

        merged_df = num_queries_df
        if sheet_name == "general_metrics":
            merged_df = merged_df.merge(avg_times_df, on=merge_columns, how="left")
        merged_df = merged_df.merge(count_top_1_df, on=merge_columns, how="left")
        merged_df = merged_df.merge(count_top_3_df, on=merge_columns, how="left")
        merged_df = merged_df.merge(count_top_5_df, on=merge_columns, how="left")
        merged_df = merged_df.merge(mrr_sum_df, on=merge_columns, how="left")

        # Compute metrics
        merged_df["Top 1"] = merged_df["count_top_1"] * 100 / merged_df["num_queries"]
        merged_df["Top 3"] = merged_df["count_top_3"] * 100 / merged_df["num_queries"]
        merged_df["Top 5"] = merged_df["count_top_5"] * 100 / merged_df["num_queries"]
        merged_df["MRR"] = merged_df["sum_ranking_pos"] / merged_df["num_queries"]

        # Drop unnecessary columns
        merged_df.drop(
            columns=["count_top_1", "count_top_3", "count_top_5", "sum_ranking_pos"],
            inplace=True,
        )

        # Write DataFrame to worksheet
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(
            dataframe_to_rows(merged_df, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True, color="FFFFFF")
                else:
                    if c_idx == len(merge_columns) + 1:
                        cell.number_format = "0"
                    else:
                        cell.number_format = "0.000"
                    if isinstance(cell.value, (int, float)):
                        cell.value = round(cell.value, 3)

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        table = Table(displayName=f"Table_{sheet_name}", ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium13",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(output_file)
    conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Analyze folkoteca database results.")
    parser.add_argument(
        "--use-global-alignment",
        action="store_true",
        default=False,
        help="Use global alignment for analysis.",
    )
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.abspath(os.path.join(script_dir, "folkoteca.db"))
    global_db_path = os.path.abspath(
        os.path.join(script_dir, "../analysis/global_folkoteca.db")
    )
    excel_filename = os.path.abspath(os.path.join(script_dir, "folkoteca_report.xlsx"))
    calculate_metrics(
        db_path, global_db_path, excel_filename, args.use_global_alignment
    )
